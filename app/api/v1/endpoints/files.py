"""
파일 업로드 엔드포인트
- 이미지: base64 인코딩 + 리사이즈
- 문서(PDF, DOCX, TXT, CSV, XLSX): 텍스트 추출

보안:
- 3중 검증: 매직 바이트(파일 시그니처) + MIME 타입 + 파일 확장자
- 파일명 sanitization (경로 순회, 특수문자, 길이 제한)
- 문서 파서 보호 (페이지/행 수 제한, 추출 텍스트 크기 제한)
- Pillow 이미지 디컴프레션 폭탄 방지
"""

import base64
import io
import logging
import os
import re
import unicodedata
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from pydantic import BaseModel

from app.core.security import get_current_user
from app.models.user import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/files", tags=["Files"])

# ─── 환경변수 설정 ────────────────────────────────────────────────────────────
MAX_UPLOAD_FILE_SIZE = int(os.environ.get("MAX_UPLOAD_FILE_SIZE", "20")) * 1024 * 1024  # MB → bytes
MAX_IMAGE_DIMENSION = int(os.environ.get("MAX_IMAGE_DIMENSION", "1280"))

# ─── 보안 상수 ────────────────────────────────────────────────────────────────
MAX_FILENAME_LENGTH = 255           # 파일명 최대 길이
MAX_PDF_PAGES = 200                 # PDF 최대 페이지 수
MAX_XLSX_ROWS = 10000               # XLSX 최대 행 수
MAX_EXTRACTED_TEXT_SIZE = 500_000    # 추출 텍스트 최대 글자수 (약 500KB)
MAX_IMAGE_PIXELS = 50_000_000       # 이미지 최대 픽셀 수 (디컴프레션 폭탄 방지, ~7000x7000)

# ─── 허용 MIME 타입 ───────────────────────────────────────────────────────────
ALLOWED_IMAGE_TYPES = {
    "image/jpeg", "image/png", "image/gif", "image/webp", "image/bmp",
}
ALLOWED_DOCUMENT_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # docx
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # xlsx
    "text/csv",
    "text/plain",
}
ALL_ALLOWED_TYPES = ALLOWED_IMAGE_TYPES | ALLOWED_DOCUMENT_TYPES

# ─── 확장자 → MIME 타입 매핑 ─────────────────────────────────────────────────
EXTENSION_TO_MIME = {
    ".jpg": {"image/jpeg"},
    ".jpeg": {"image/jpeg"},
    ".png": {"image/png"},
    ".gif": {"image/gif"},
    ".webp": {"image/webp"},
    ".bmp": {"image/bmp"},
    ".pdf": {"application/pdf"},
    ".docx": {"application/vnd.openxmlformats-officedocument.wordprocessingml.document"},
    ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    ".csv": {"text/csv", "text/plain"},
    ".txt": {"text/plain", "text/csv"},
}
ALLOWED_EXTENSIONS = set(EXTENSION_TO_MIME.keys())

# ─── 매직 바이트 (파일 시그니처) ──────────────────────────────────────────────
# 각 파일 포맷의 첫 바이트를 검증하여 위변조 방지
MAGIC_SIGNATURES: dict[str, list[tuple[bytes, int]]] = {
    # (시그니처 바이트, 오프셋) 리스트
    "image/jpeg": [(b"\xff\xd8\xff", 0)],
    "image/png": [(b"\x89PNG\r\n\x1a\n", 0)],
    "image/gif": [(b"GIF87a", 0), (b"GIF89a", 0)],
    "image/webp": [(b"RIFF", 0)],  # RIFF 컨테이너, 추가로 WEBP 확인
    "image/bmp": [(b"BM", 0)],
    "application/pdf": [(b"%PDF", 0)],
    # DOCX/XLSX는 ZIP 기반 (PK 시그니처)
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": [(b"PK\x03\x04", 0)],
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": [(b"PK\x03\x04", 0)],
    # CSV/TXT는 매직 바이트 없음 — 별도 검증
}


def _check_magic_bytes(file_bytes: bytes, expected_mime: str) -> bool:
    """파일의 매직 바이트(시그니처)를 검증하여 실제 파일 포맷 확인"""
    if expected_mime in ("text/csv", "text/plain"):
        # 텍스트 파일: NULL 바이트가 없어야 함 (바이너리 파일 차단)
        # 처음 8KB만 검사 (성능)
        sample = file_bytes[:8192]
        if b"\x00" in sample:
            return False
        return True

    signatures = MAGIC_SIGNATURES.get(expected_mime)
    if signatures is None:
        return False

    for sig_bytes, offset in signatures:
        if file_bytes[offset:offset + len(sig_bytes)] == sig_bytes:
            # WebP 추가 검증: RIFF 컨테이너 내부에 "WEBP" 문자열 확인
            if expected_mime == "image/webp":
                return file_bytes[8:12] == b"WEBP"
            return True

    return False


def _sanitize_filename(filename: str) -> str:
    """
    파일명 보안 처리:
    - 경로 순회 문자 제거 (../ ..\\ 등)
    - 디렉토리 구분자 제거
    - 제어 문자 및 위험 특수문자 제거
    - 유니코드 정규화 (NFC)
    - 길이 제한
    - 빈 파일명 방지
    """
    if not filename:
        return "unknown"

    # 유니코드 정규화
    filename = unicodedata.normalize("NFC", filename)

    # 경로 구분자 제거 → 파일명만 추출
    filename = filename.replace("\\", "/")
    filename = filename.split("/")[-1]

    # 제어 문자 제거 (U+0000 ~ U+001F, U+007F ~ U+009F)
    filename = re.sub(r"[\x00-\x1f\x7f-\x9f]", "", filename)

    # 위험 특수문자 제거 (HTML/JS 인젝션 방지)
    filename = re.sub(r'[<>:"/\\|?*\x00]', "_", filename)

    # 연속 점/공백 제거 (.. 공격 방지)
    filename = re.sub(r"\.{2,}", ".", filename)
    filename = filename.strip(". ")

    # 길이 제한
    if len(filename) > MAX_FILENAME_LENGTH:
        # 확장자 보존
        name_part, _, ext = filename.rpartition(".")
        if ext:
            filename = name_part[:MAX_FILENAME_LENGTH - len(ext) - 1] + "." + ext
        else:
            filename = filename[:MAX_FILENAME_LENGTH]

    # 빈 파일명 방지
    if not filename or filename == ".":
        filename = "unknown"

    return filename


def _get_file_extension(filename: str) -> str:
    """파일명에서 소문자 확장자 추출"""
    if "." not in filename:
        return ""
    return "." + filename.rsplit(".", 1)[-1].lower()


def _validate_file(file_bytes: bytes, content_type: str, filename: str) -> str:
    """
    3중 파일 검증: MIME 타입 + 확장자 + 매직 바이트
    검증 통과 시 검증된 MIME 타입 반환, 실패 시 HTTPException 발생
    """
    # 1단계: MIME 타입 검증
    if content_type not in ALL_ALLOWED_TYPES:
        logger.warning("파일 업로드 거부 — 허용되지 않는 MIME 타입: %s (파일: %s)", content_type, filename)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"지원하지 않는 파일 형식입니다: {content_type}. "
                   f"지원 형식: 이미지(JPEG, PNG, GIF, WebP, BMP), 문서(PDF, DOCX, XLSX, CSV, TXT)",
        )

    # 2단계: 확장자 검증 (MIME 타입과 확장자 일치 여부)
    ext = _get_file_extension(filename)
    if ext:
        allowed_mimes_for_ext = EXTENSION_TO_MIME.get(ext)
        if allowed_mimes_for_ext is None:
            logger.warning("파일 업로드 거부 — 허용되지 않는 확장자: %s (MIME: %s, 파일: %s)", ext, content_type, filename)
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"허용되지 않는 파일 확장자입니다: {ext}",
            )
        if content_type not in allowed_mimes_for_ext:
            logger.warning(
                "파일 업로드 거부 — MIME/확장자 불일치: MIME=%s, 확장자=%s (파일: %s)",
                content_type, ext, filename,
            )
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"파일 확장자({ext})와 파일 형식({content_type})이 일치하지 않습니다.",
            )

    # 3단계: 매직 바이트(파일 시그니처) 검증
    if not _check_magic_bytes(file_bytes, content_type):
        logger.warning(
            "파일 업로드 거부 — 매직 바이트 불일치: MIME=%s (파일: %s)",
            content_type, filename,
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="파일 내용이 선언된 형식과 일치하지 않습니다. 파일이 손상되었거나 위변조된 것 같습니다.",
        )

    return content_type


def _truncate_text(text: str, max_length: int = MAX_EXTRACTED_TEXT_SIZE) -> str:
    """추출된 텍스트가 최대 길이를 초과하면 잘라냄"""
    if len(text) > max_length:
        return text[:max_length] + f"\n\n... (텍스트가 {max_length:,}자를 초과하여 잘렸습니다)"
    return text


class FileUploadResponse(BaseModel):
    """파일 업로드 응답"""
    type: str  # "image" | "document"
    content: str  # base64 data URI (이미지) 또는 추출된 텍스트 (문서)
    filename: str
    mime_type: str


def _resize_image(image_bytes: bytes, content_type: str) -> tuple[str, str]:
    """
    이미지를 MAX_IMAGE_DIMENSION 이내로 리사이즈하고 base64 data URI 반환

    보안:
    - Pillow MAX_IMAGE_PIXELS로 디컴프레션 폭탄 방지
    - 이미지 포맷 재검증 (Pillow가 실제 포맷 확인)
    """
    from PIL import Image, ImageOps

    # 디컴프레션 폭탄 방지 (예: 1x1 헤더에 거대한 픽셀 데이터)
    Image.MAX_IMAGE_PIXELS = MAX_IMAGE_PIXELS

    try:
        img = Image.open(io.BytesIO(image_bytes))
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="이미지 파일을 열 수 없습니다. 파일이 손상되었을 수 있습니다.",
        ) from e

    # Pillow가 인식한 포맷 재검증
    pillow_format = (img.format or "").upper()
    format_mime_map = {"JPEG": "image/jpeg", "PNG": "image/png", "GIF": "image/gif",
                       "WEBP": "image/webp", "BMP": "image/bmp"}
    detected_mime = format_mime_map.get(pillow_format)
    if detected_mime and detected_mime != content_type:
        # BMP → JPEG 등 변환 케이스는 허용하되, 비이미지 포맷은 차단
        if detected_mime not in ALLOWED_IMAGE_TYPES:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="이미지 파일의 실제 포맷이 선언된 형식과 다릅니다.",
            )

    # EXIF 회전 정보 적용
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass

    # 리사이즈 (가로/세로 중 큰 쪽이 MAX_IMAGE_DIMENSION 초과 시)
    max_dim = MAX_IMAGE_DIMENSION
    if img.width > max_dim or img.height > max_dim:
        ratio = min(max_dim / img.width, max_dim / img.height)
        new_size = (int(img.width * ratio), int(img.height * ratio))
        img = img.resize(new_size, Image.LANCZOS)

    # RGBA → RGB 변환 (JPEG 호환)
    output_format = "JPEG"
    output_mime = "image/jpeg"
    if content_type == "image/png":
        output_format = "PNG"
        output_mime = "image/png"
    elif content_type == "image/webp":
        output_format = "WEBP"
        output_mime = "image/webp"
    elif content_type == "image/gif":
        output_format = "GIF"
        output_mime = "image/gif"

    if output_format == "JPEG" and img.mode in ("RGBA", "LA", "P"):
        img = img.convert("RGB")

    buf = io.BytesIO()
    img.save(buf, format=output_format, quality=85, optimize=True)
    encoded = base64.b64encode(buf.getvalue()).decode("utf-8")

    return f"data:{output_mime};base64,{encoded}", output_mime


def _extract_pdf_text(file_bytes: bytes) -> str:
    """
    PDF에서 텍스트 추출

    보안: 최대 페이지 수 제한 (악성 PDF의 무한 페이지 공격 방지)
    """
    import pdfplumber

    text_parts = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            total_pages = len(pdf.pages)
            if total_pages > MAX_PDF_PAGES:
                logger.warning("PDF 페이지 수 초과: %d (최대 %d)", total_pages, MAX_PDF_PAGES)
                text_parts.append(f"(전체 {total_pages}페이지 중 처음 {MAX_PDF_PAGES}페이지만 추출)")

            for i, page in enumerate(pdf.pages[:MAX_PDF_PAGES], 1):
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(f"[페이지 {i}]\n{page_text}")
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="PDF 파일을 처리할 수 없습니다. 파일이 손상되었거나 지원되지 않는 형식입니다.",
        ) from e

    result = "\n\n".join(text_parts) if text_parts else "(PDF에서 텍스트를 추출할 수 없습니다.)"
    return _truncate_text(result)


def _extract_docx_text(file_bytes: bytes) -> str:
    """
    DOCX에서 텍스트 추출

    보안: 추출 텍스트 크기 제한
    """
    from docx import Document

    try:
        doc = Document(io.BytesIO(file_bytes))
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="DOCX 파일을 처리할 수 없습니다. 파일이 손상되었거나 지원되지 않는 형식입니다.",
        ) from e

    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    result = "\n\n".join(paragraphs) if paragraphs else "(DOCX에서 텍스트를 추출할 수 없습니다.)"
    return _truncate_text(result)


def _extract_xlsx_text(file_bytes: bytes) -> str:
    """
    XLSX에서 텍스트 추출 (각 시트를 테이블 형태로)

    보안: 최대 행 수 제한 (거대 스프레드시트의 메모리 과다 사용 방지)
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="XLSX 파일을 처리할 수 없습니다. 파일이 손상되었거나 지원되지 않는 형식입니다.",
        ) from e

    parts = []
    total_rows = 0
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if total_rows >= MAX_XLSX_ROWS:
                    rows.append(f"... (최대 {MAX_XLSX_ROWS:,}행 제한으로 잘림)")
                    break
                row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                rows.append(row_str)
                total_rows += 1
            if rows:
                parts.append(f"[시트: {sheet_name}]\n" + "\n".join(rows))
            if total_rows >= MAX_XLSX_ROWS:
                break
    finally:
        wb.close()

    result = "\n\n".join(parts) if parts else "(XLSX에서 데이터를 추출할 수 없습니다.)"
    return _truncate_text(result)


def _extract_csv_text(file_bytes: bytes) -> str:
    """
    CSV 텍스트 읽기

    보안: 텍스트 크기 제한
    """
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode("euc-kr")
        except UnicodeDecodeError:
            text = file_bytes.decode("utf-8", errors="replace")
    return _truncate_text(text)


@router.post("/upload", response_model=FileUploadResponse)
async def upload_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """
    파일 업로드 및 처리

    - 이미지: 리사이즈 후 base64 data URI로 반환
    - 문서: 텍스트 추출 후 반환

    보안:
    - 3중 검증: 매직 바이트 + MIME 타입 + 확장자
    - 파일명 sanitization
    - 이미지 디컴프레션 폭탄 방지
    - 문서 파서 크기 제한
    """
    # 파일명 sanitization (경로 순회, 특수문자, 길이 제한)
    raw_filename = file.filename or "unknown"
    filename = _sanitize_filename(raw_filename)
    content_type = file.content_type or ""

    # 파일 크기 검증 (바이트 읽기 전 빈 파일 체크)
    file_bytes = await file.read()

    if len(file_bytes) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="빈 파일은 업로드할 수 없습니다.",
        )

    if len(file_bytes) > MAX_UPLOAD_FILE_SIZE:
        max_mb = MAX_UPLOAD_FILE_SIZE // (1024 * 1024)
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"파일 크기가 {max_mb}MB를 초과합니다.",
        )

    # 3중 검증: MIME 타입 + 확장자 + 매직 바이트
    validated_type = _validate_file(file_bytes, content_type, filename)

    logger.info(
        "파일 업로드 — 사용자: %s, 파일: %s, 크기: %s bytes, 타입: %s",
        current_user.username, filename, len(file_bytes), validated_type,
    )

    try:
        # 이미지 처리
        if validated_type in ALLOWED_IMAGE_TYPES:
            data_uri, output_mime = _resize_image(file_bytes, validated_type)
            return FileUploadResponse(
                type="image",
                content=data_uri,
                filename=filename,
                mime_type=output_mime,
            )

        # 문서 처리
        extracted_text: Optional[str] = None

        if validated_type == "application/pdf":
            extracted_text = _extract_pdf_text(file_bytes)
        elif validated_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            extracted_text = _extract_docx_text(file_bytes)
        elif validated_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            extracted_text = _extract_xlsx_text(file_bytes)
        elif validated_type == "text/csv":
            extracted_text = _extract_csv_text(file_bytes)
        elif validated_type == "text/plain":
            try:
                text = file_bytes.decode("utf-8")
            except UnicodeDecodeError:
                text = file_bytes.decode("utf-8", errors="replace")
            extracted_text = _truncate_text(text)

        if extracted_text is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="파일에서 내용을 추출할 수 없습니다.",
            )

        return FileUploadResponse(
            type="document",
            content=extracted_text,
            filename=filename,
            mime_type=validated_type,
        )

    except HTTPException:
        raise
    except Exception as e:
        # 내부 에러 메시지를 사용자에게 그대로 노출하지 않음 (보안)
        logger.error("파일 처리 중 오류 — 파일: %s, 에러: %s", filename, str(e), exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="파일 처리 중 오류가 발생했습니다. 다른 파일로 시도해 주세요.",
        ) from e
